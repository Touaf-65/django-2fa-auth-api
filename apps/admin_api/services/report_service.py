"""
Service pour la génération de rapports
"""
import json
import csv
import os
from datetime import datetime, timedelta
from django.conf import settings
from django.db.models import Count, Q
from django.contrib.auth import get_user_model
from django.utils import timezone
from apps.admin_api.models import ReportTemplate, ScheduledReport, ReportExecution

User = get_user_model()


class ReportService:
    """Service pour la génération de rapports"""
    
    def __init__(self):
        self.reports_dir = getattr(settings, 'REPORTS_DIR', 'reports')
        self._ensure_reports_dir()
    
    def _ensure_reports_dir(self):
        """S'assure que le répertoire des rapports existe"""
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def generate_report(self, template, parameters=None):
        """Génère un rapport basé sur un template"""
        parameters = parameters or {}
        
        try:
            # Récupère les données selon le type de rapport
            data = self._get_report_data(template, parameters)
            
            # Génère le fichier selon le format
            file_path = self._generate_file(template, data)
            
            return {
                'success': True,
                'file_path': file_path,
                'record_count': len(data) if isinstance(data, list) else 1,
                'data': data
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def _get_report_data(self, template, parameters):
        """Récupère les données pour le rapport"""
        report_type = template.report_type
        
        if report_type == 'users':
            return self._get_users_data(parameters)
        elif report_type == 'activity':
            return self._get_activity_data(parameters)
        elif report_type == 'security':
            return self._get_security_data(parameters)
        elif report_type == 'performance':
            return self._get_performance_data(parameters)
        elif report_type == 'system':
            return self._get_system_data(parameters)
        elif report_type == 'errors':
            return self._get_errors_data(parameters)
        elif report_type == 'api_usage':
            return self._get_api_usage_data(parameters)
        else:
            return []
    
    def _get_users_data(self, parameters):
        """Données des utilisateurs"""
        from apps.authentication.models import User
        from apps.users.models import UserProfile
        
        queryset = User.objects.all()
        
        # Applique les filtres
        if parameters.get('date_from'):
            queryset = queryset.filter(date_joined__gte=parameters['date_from'])
        if parameters.get('date_to'):
            queryset = queryset.filter(date_joined__lte=parameters['date_to'])
        if parameters.get('is_active') is not None:
            queryset = queryset.filter(is_active=parameters['is_active'])
        
        users_data = []
        for user in queryset:
            try:
                profile = user.userprofile
            except:
                profile = None
            
            users_data.append({
                'id': user.id,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'is_active': user.is_active,
                'is_staff': user.is_staff,
                'date_joined': user.date_joined.isoformat() if user.date_joined else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'profile_completed': profile is not None,
            })
        
        return users_data
    
    def _get_activity_data(self, parameters):
        """Données d'activité"""
        from apps.admin_api.models import AdminAction, AdminLog
        
        actions_data = []
        logs_data = []
        
        # Actions d'administration
        actions = AdminAction.objects.all()
        if parameters.get('date_from'):
            actions = actions.filter(created_at__gte=parameters['date_from'])
        if parameters.get('date_to'):
            actions = actions.filter(created_at__lte=parameters['date_to'])
        
        for action in actions:
            actions_data.append({
                'id': action.id,
                'action_type': action.action_type,
                'title': action.title,
                'status': action.status,
                'admin_user': action.admin_user.email,
                'created_at': action.created_at.isoformat(),
                'duration': str(action.duration) if action.duration else None,
            })
        
        # Logs d'administration
        logs = AdminLog.objects.all()
        if parameters.get('date_from'):
            logs = logs.filter(created_at__gte=parameters['date_from'])
        if parameters.get('date_to'):
            logs = logs.filter(created_at__lte=parameters['date_to'])
        
        for log in logs:
            logs_data.append({
                'id': log.id,
                'level': log.level,
                'action': log.action,
                'message': log.message,
                'admin_user': log.admin_user.email,
                'created_at': log.created_at.isoformat(),
            })
        
        return {
            'actions': actions_data,
            'logs': logs_data
        }
    
    def _get_security_data(self, parameters):
        """Données de sécurité"""
        from apps.security.models import LoginAttempt, SecurityEvent, IPBlock
        
        security_data = {
            'login_attempts': [],
            'security_events': [],
            'blocked_ips': []
        }
        
        # Tentatives de connexion
        login_attempts = LoginAttempt.objects.all()
        if parameters.get('date_from'):
            login_attempts = login_attempts.filter(attempted_at__gte=parameters['date_from'])
        if parameters.get('date_to'):
            login_attempts = login_attempts.filter(attempted_at__lte=parameters['date_to'])
        
        for attempt in login_attempts:
            security_data['login_attempts'].append({
                'id': attempt.id,
                'email': attempt.email,
                'ip_address': attempt.ip_address,
                'success': attempt.success,
                'attempted_at': attempt.attempted_at.isoformat(),
            })
        
        # Événements de sécurité
        security_events = SecurityEvent.objects.all()
        if parameters.get('date_from'):
            security_events = security_events.filter(created_at__gte=parameters['date_from'])
        if parameters.get('date_to'):
            security_events = security_events.filter(created_at__lte=parameters['date_to'])
        
        for event in security_events:
            security_data['security_events'].append({
                'id': event.id,
                'event_type': event.event_type,
                'severity': event.severity,
                'description': event.description,
                'ip_address': event.ip_address,
                'created_at': event.created_at.isoformat(),
            })
        
        return security_data
    
    def _get_performance_data(self, parameters):
        """Données de performance"""
        import psutil
        
        return {
            'cpu_percent': psutil.cpu_percent(interval=1),
            'memory_percent': psutil.virtual_memory().percent,
            'disk_percent': psutil.disk_usage('/').percent,
            'timestamp': timezone.now().isoformat(),
        }
    
    def _get_system_data(self, parameters):
        """Données système"""
        import platform
        import sys
        
        return {
            'platform': {
                'system': platform.system(),
                'release': platform.release(),
                'version': platform.version(),
                'machine': platform.machine(),
            },
            'python': {
                'version': sys.version,
                'executable': sys.executable,
            },
            'django': {
                'version': settings.DJANGO_VERSION,
                'debug': settings.DEBUG,
            },
            'timestamp': timezone.now().isoformat(),
        }
    
    def _get_errors_data(self, parameters):
        """Données d'erreurs"""
        from apps.admin_api.models import AdminLog
        
        error_logs = AdminLog.objects.filter(level__in=['error', 'critical'])
        if parameters.get('date_from'):
            error_logs = error_logs.filter(created_at__gte=parameters['date_from'])
        if parameters.get('date_to'):
            error_logs = error_logs.filter(created_at__lte=parameters['date_to'])
        
        errors_data = []
        for log in error_logs:
            errors_data.append({
                'id': log.id,
                'level': log.level,
                'action': log.action,
                'message': log.message,
                'admin_user': log.admin_user.email,
                'created_at': log.created_at.isoformat(),
            })
        
        return errors_data
    
    def _get_api_usage_data(self, parameters):
        """Données d'utilisation de l'API"""
        # Ici vous pouvez implémenter la logique pour suivre l'utilisation de l'API
        # Par exemple, en utilisant des middlewares ou des logs
        return {
            'message': 'API usage tracking not implemented yet',
            'timestamp': timezone.now().isoformat(),
        }
    
    def _generate_file(self, template, data):
        """Génère le fichier de rapport"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{template.name}_{timestamp}.{template.format}"
        file_path = os.path.join(self.reports_dir, filename)
        
        if template.format == 'json':
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        elif template.format == 'csv':
            if isinstance(data, list) and data:
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            else:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write("No data available")
        
        elif template.format == 'xlsx':
            # Pour Excel, vous devrez installer openpyxl
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = template.name
                
                if isinstance(data, list) and data:
                    # Écriture des en-têtes
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Écriture des données
                    for row, item in enumerate(data, 2):
                        for col, key in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=item.get(key, ''))
                
                wb.save(file_path)
            except ImportError:
                raise Exception("openpyxl is required for Excel export")
        
        return file_path
    
    def execute_scheduled_report(self, scheduled_report):
        """Exécute un rapport programmé"""
        execution = ReportExecution.objects.create(
            scheduled_report=scheduled_report,
            template=scheduled_report.template,
            status='pending'
        )
        
        try:
            execution.start()
            
            # Génère le rapport
            result = self.generate_report(
                scheduled_report.template,
                scheduled_report.template.filters
            )
            
            if result['success']:
                execution.complete(
                    file_path=result['file_path'],
                    record_count=result['record_count']
                )
                
                # Met à jour la prochaine exécution
                scheduled_report.last_run = timezone.now()
                scheduled_report.next_run = scheduled_report.calculate_next_run()
                scheduled_report.save()
                
                return execution
            else:
                execution.fail(result['error'])
                return execution
                
        except Exception as e:
            execution.fail(str(e))
            return execution

